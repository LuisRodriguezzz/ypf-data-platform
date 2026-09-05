"""Un XLSX chico con el mismo encabezado que la planilla real de reservas.

Se genera con openpyxl en vez de guardar un binario en el repo: el test dice cual es el
layout que el parser tiene que soportar, y se lee igual que el codigo.
"""

from __future__ import annotations

import io

import openpyxl
import pytest

# Cada bloque ocupa 8 columnas: 6 de reservas (comprobadas, probables, posibles por PET y
# GAS) y 2 de recursos contingentes, que no se subdividen por certeza.
BLOQUES = (("CONVENCIONAL", 6), ("NO CONVENCIONAL", 14), ("CONVENCIONAL + NO CONVENCIONAL", 22))
CERTEZAS = ("COMPROBADAS", "PROBABLES", "POSIBLES")

IDENTIDAD = ("OPERADOR", "CUENCA", "PROVINCIA", "CONCESIÓN O PERMISO", "YACIMIENTO")

OPERADORES = (
    ("YPF S.A.", "NEUQUINA", "Neuquén", "LOMA CAMPANA", "LOMA CAMPANA"),
    ("PAN AMERICAN ENERGY SL", "GOLFO SAN JORGE", "Chubut", "CERRO DRAGON", "ANTICLINAL FUNES"),
)

PRIMERA_FILA_DATOS = 8


def celda_esperada(fila: int, columna: int) -> float:
    """Valor deterministico de la celda, para poder afirmar sobre valores puntuales.

    El `.5` evita que openpyxl devuelva un entero: interesa que el numero llegue a bronze
    como texto sin perder decimales.
    """
    return fila * 100 + columna + 0.5


def _escribir_encabezado(hoja, titulo: str) -> None:
    hoja.cell(row=1, column=6, value=titulo)
    hoja.merge_cells(start_row=1, start_column=6, end_row=2, end_column=29)
    for nombre, base in BLOQUES:
        hoja.cell(row=3, column=base, value=nombre)
        hoja.merge_cells(start_row=3, start_column=base, end_row=3, end_column=base + 7)
        hoja.cell(row=4, column=base, value="RESERVAS")
        hoja.merge_cells(start_row=4, start_column=base, end_row=4, end_column=base + 5)
        for indice, certeza in enumerate(CERTEZAS):
            columna = base + indice * 2
            hoja.cell(row=5, column=columna, value=certeza)
            hoja.merge_cells(start_row=5, start_column=columna, end_row=5, end_column=columna + 1)
        # Rango fusionado a lo alto: el bloque no se subdivide por certeza.
        hoja.cell(row=4, column=base + 6, value="RECURSOS CONTINGENTES")
        hoja.merge_cells(start_row=4, start_column=base + 6, end_row=5, end_column=base + 7)
        for offset in range(8):
            columna = base + offset
            es_petroleo = offset % 2 == 0
            hoja.cell(row=6, column=columna, value="PET" if es_petroleo else "GAS")
            hoja.cell(row=7, column=columna, value="(Mm3)" if es_petroleo else "(MMm3)")
    for indice, nombre in enumerate(IDENTIDAD, start=1):
        hoja.cell(row=7, column=indice, value=nombre)


def _escribir_datos(hoja) -> None:
    for offset, operador in enumerate(OPERADORES):
        fila = PRIMERA_FILA_DATOS + offset
        for indice, valor in enumerate(operador, start=1):
            hoja.cell(row=fila, column=indice, value=valor)
        for columna in range(6, 30):
            hoja.cell(row=fila, column=columna, value=celda_esperada(fila, columna))
    # Fila separadora y fila de cierre, como en la planilla real.
    hoja.cell(row=PRIMERA_FILA_DATOS + len(OPERADORES) + 1, column=1, value="TOTAL")
    hoja.cell(row=PRIMERA_FILA_DATOS + len(OPERADORES) + 1, column=6, value=99999.0)


@pytest.fixture
def planilla() -> bytes:
    """XLSX con las dos hojas, dos operadores y el encabezado de 7 filas fusionado."""
    libro = openpyxl.Workbook()
    libro.remove(libro.active)
    for titulo, encabezado in (
        ("fin de concesión", "RESERVAS Y RECURSOS AL 31/12/2024 - HASTA EL FIN DE LA CONCESIÓN"),
        ("fin de vida util", "RESERVAS Y RECURSOS AL 31/12/2024 - HASTA EL FIN DE LA VIDA ÚTIL"),
    ):
        hoja = libro.create_sheet(titulo)
        _escribir_encabezado(hoja, encabezado)
        _escribir_datos(hoja)
    buffer = io.BytesIO()
    libro.save(buffer)
    return buffer.getvalue()
