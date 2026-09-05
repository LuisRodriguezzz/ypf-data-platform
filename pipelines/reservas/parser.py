"""Parser del XLSX anual de reservas: de la planilla ancha a filas largas.

La planilla no es una tabla: es un cuadro de doble entrada con cuatro niveles de
encabezado fusionados (tipo de recurso, categoria, certeza, fluido) sobre cinco columnas
de identificacion. Este modulo lo aplana a una fila por celda de valor, que es la forma
que puede guardar una tabla Iceberg y consultar SQL.

Son funciones puras sobre bytes o rutas: no tocan S3 ni el catalogo, asi que se testean
sin infraestructura.
"""

from __future__ import annotations

import io
import logging
import re
import unicodedata
import zipfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)

# Columnas de la fila larga, en el orden en que quedan en bronze.
LONG_COLUMNS = (
    "operador",
    "cuenca",
    "provincia",
    "concesion",
    "yacimiento",
    "hoja",
    "tipo_recurso",
    "categoria",
    "certeza",
    "fluido",
    "unidad",
    "valor",
    "anio_corte",
)

# Rotulos de la fila de encabezado -> nombre de la columna larga.
IDENTITY_LABELS = {
    "OPERADOR": "operador",
    "CUENCA": "cuenca",
    "PROVINCIA": "provincia",
    "CONCESION O PERMISO": "concesion",
    "YACIMIENTO": "yacimiento",
}

# Vocabulario de cada nivel del encabezado. Sirve para dos cosas: traducir el rotulo y
# reconocer que fila del encabezado es cual sin depender de su posicion.
FLUIDOS = {"PET": "petroleo", "GAS": "gas"}
TIPOS_RECURSO = {"CONVENCIONAL": "convencional", "NO CONVENCIONAL": "no_convencional"}
CATEGORIAS = {"RESERVAS": "reservas", "RECURSOS CONTINGENTES": "recursos_contingentes"}
CERTEZAS = {"COMPROBADAS": "comprobadas", "PROBABLES": "probables", "POSIBLES": "posibles"}

# Los recursos contingentes no se subdividen en certeza. Se marca con un valor y no con
# vacio porque `certeza` es parte de la clave primaria, y una clave con nulos no es clave.
SIN_CERTEZA = "no_aplica"

# Filas de cierre del cuadro: la suma de la columna, no un yacimiento.
TOTAL_LABELS = {"TOTAL", "TOTALES", "TOTAL GENERAL"}

HEADER_SEARCH_ROWS = 20


@dataclass(frozen=True)
class ValueColumn:
    """Una columna de valores del cuadro con los cuatro niveles que la describen."""

    column: int
    tipo_recurso: str
    categoria: str
    certeza: str
    fluido: str
    unidad: str


@dataclass(frozen=True)
class SheetLayout:
    """Donde estan los rotulos y los datos de una hoja."""

    header_row: int
    identity: dict[str, int]
    values: tuple[ValueColumn, ...]


@dataclass
class ParseResult:
    """Filas largas de un archivo mas lo que se descarto al leerlo."""

    rows: list[dict[str, str]] = field(default_factory=list)
    sheets: list[str] = field(default_factory=list)
    skipped_totals: int = 0


def normalize(value: Any) -> str:
    """Texto sin acentos, en mayusculas y con espacios colapsados; '' si la celda esta vacia."""
    if value is None:
        return ""
    plano = unicodedata.normalize("NFKD", str(value)).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"\s+", " ", plano).strip().upper()


def text(value: Any) -> str:
    """Valor de celda como texto para bronze, que no tipa nada."""
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value)


def anio_from_name(name: str) -> int:
    """Anio de corte del nombre del archivo (`reservas al 31-12-2024.xlsx` -> 2024)."""
    anios = re.findall(r"(?:19|20)\d{2}", name)
    if not anios:
        raise ValueError(f"no se puede deducir el anio de corte de {name!r}")
    # El ultimo: el nombre trae la fecha completa (31-12-2024) y el anio va al final.
    return int(anios[-1])


def sheet_kind(title: str) -> str | None:
    """`fin de concesión` -> `fin_concesion`. None si la hoja no es una de las dos esperadas.

    Se busca por palabra y no por igualdad: 2020 rotula las hojas `Fin Concesion` y
    `Fin de vida útil`, y 2021-2024 `fin de concesión` y `fin de vida util`.
    """
    plano = normalize(title)
    if "VIDA" in plano:
        return "fin_vida_util"
    if "CONCESION" in plano:
        return "fin_concesion"
    return None


def expand_merges(sheet: Worksheet) -> dict[tuple[int, int], Any]:
    """Valores de los rangos fusionados propagados hacia la derecha (no hacia abajo).

    openpyxl deja el valor solo en la celda de arriba a la izquierda del rango. Se propaga
    a lo ancho porque un rotulo fusionado horizontalmente encabeza todas esas columnas;
    no se propaga a lo alto porque un rango vertical (RECURSOS CONTINGENTES ocupa la fila
    de categoria y la de certeza) significa que ese bloque no se subdivide, no que la
    subdivision se llame igual que el bloque.
    """
    valores: dict[tuple[int, int], Any] = {}
    for rango in sheet.merged_cells.ranges:
        origen = sheet.cell(row=rango.min_row, column=rango.min_col).value
        for column in range(rango.min_col, rango.max_col + 1):
            valores[(rango.min_row, column)] = origen
    return valores


class Grid:
    """La hoja leida con los rangos fusionados ya resueltos."""

    def __init__(self, sheet: Worksheet) -> None:
        self.sheet = sheet
        self.merged = expand_merges(sheet)
        self.max_row = sheet.max_row or 0
        self.max_column = sheet.max_column or 0

    def value(self, row: int, column: int) -> Any:
        if (row, column) in self.merged:
            return self.merged[(row, column)]
        return self.sheet.cell(row=row, column=column).value

    def row_labels(self, row: int) -> set[str]:
        return {normalize(self.value(row, column)) for column in range(1, self.max_column + 1)}


def find_header_row(grid: Grid) -> int:
    """Fila que trae `OPERADOR`: es la que da los nombres de las columnas de identificacion."""
    for row in range(1, min(grid.max_row, HEADER_SEARCH_ROWS) + 1):
        if "OPERADOR" in grid.row_labels(row):
            return row
    raise ValueError("la hoja no tiene una fila de encabezado con OPERADOR")


def find_label_rows(grid: Grid, header_row: int) -> dict[str, int]:
    """Ubica cada nivel del encabezado por su vocabulario, no por su distancia al header.

    Asi un titulo de mas o de menos en alguna edicion del archivo no corre los niveles.
    """
    niveles = (
        ("fluido", FLUIDOS),
        ("certeza", CERTEZAS),
        ("categoria", CATEGORIAS),
        ("tipo_recurso", TIPOS_RECURSO),
    )
    filas: dict[str, int] = {}
    for row in range(header_row - 1, 0, -1):
        etiquetas = grid.row_labels(row)
        for nivel, vocabulario in niveles:
            if nivel not in filas and etiquetas & set(vocabulario):
                filas[nivel] = row
    faltantes = [nivel for nivel, _ in niveles if nivel not in filas]
    if faltantes:
        raise ValueError(f"el encabezado no tiene las filas de {faltantes}")
    return filas


def unit_of(raw: Any) -> str:
    """`(Mm3)` -> `Mm3`; la unidad esta en la fila de encabezado, entre parentesis."""
    return text(raw).strip("() ")


def read_layout(grid: Grid) -> SheetLayout:
    """Columnas de identificacion y de valores, con los cuatro niveles ya traducidos."""
    header_row = find_header_row(grid)
    label_rows = find_label_rows(grid, header_row)

    identity: dict[str, int] = {}
    values: list[ValueColumn] = []
    for column in range(1, grid.max_column + 1):
        rotulo = normalize(grid.value(header_row, column))
        if rotulo in IDENTITY_LABELS:
            identity[IDENTITY_LABELS[rotulo]] = column
            continue
        fluido = FLUIDOS.get(normalize(grid.value(label_rows["fluido"], column)))
        if not fluido:
            continue
        # `CONVENCIONAL + NO CONVENCIONAL` es la suma de los otros dos bloques: es un total
        # derivable, no un dato nuevo, y guardarlo romperia la unicidad de la clave.
        tipo = TIPOS_RECURSO.get(normalize(grid.value(label_rows["tipo_recurso"], column)))
        if not tipo:
            continue
        values.append(
            ValueColumn(
                column=column,
                tipo_recurso=tipo,
                categoria=CATEGORIAS.get(
                    normalize(grid.value(label_rows["categoria"], column)), ""
                ),
                certeza=CERTEZAS.get(
                    normalize(grid.value(label_rows["certeza"], column)), SIN_CERTEZA
                ),
                fluido=fluido,
                unidad=unit_of(grid.value(header_row, column)),
            )
        )

    faltantes = sorted(set(IDENTITY_LABELS.values()) - set(identity))
    if faltantes:
        raise ValueError(f"faltan columnas de identificacion: {faltantes}")
    if not values:
        raise ValueError("la hoja no tiene columnas de valores PET/GAS")
    return SheetLayout(header_row=header_row, identity=identity, values=tuple(values))


def is_total_row(identity: dict[str, str]) -> bool:
    """Fila de cierre del cuadro: dice TOTAL en el operador y no identifica un yacimiento."""
    return normalize(identity.get("operador")) in TOTAL_LABELS


def is_empty_row(identity: dict[str, str]) -> bool:
    """Fila separadora: ninguna columna de identificacion tiene texto."""
    return not any(identity.values())


def parse_sheet(grid: Grid, hoja: str, anio_corte: int) -> ParseResult:
    """Filas largas de una hoja: una por celda de valor del cuadro."""
    layout = read_layout(grid)
    result = ParseResult(sheets=[hoja])
    for row in range(layout.header_row + 1, grid.max_row + 1):
        identity = {name: text(grid.value(row, col)) for name, col in layout.identity.items()}
        if is_empty_row(identity):
            continue
        if is_total_row(identity):
            result.skipped_totals += 1
            continue
        for value in layout.values:
            result.rows.append(
                {
                    **identity,
                    "hoja": hoja,
                    "tipo_recurso": value.tipo_recurso,
                    "categoria": value.categoria,
                    "certeza": value.certeza,
                    "fluido": value.fluido,
                    "unidad": value.unidad,
                    "valor": text(grid.value(row, value.column)),
                    "anio_corte": str(anio_corte),
                }
            )
    return result


def parse_bytes(data: bytes, anio_corte: int) -> ParseResult:
    """Filas largas de las dos hojas del XLSX, ya en memoria."""
    workbook = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    try:
        result = ParseResult()
        for sheet in workbook.worksheets:
            hoja = sheet_kind(sheet.title)
            if hoja is None:
                logger.warning("hoja ignorada, no es concesion ni vida util: %r", sheet.title)
                continue
            parcial = parse_sheet(Grid(sheet), hoja, anio_corte)
            result.rows.extend(parcial.rows)
            result.sheets.extend(parcial.sheets)
            result.skipped_totals += parcial.skipped_totals
        return result
    finally:
        workbook.close()


def parse_file(path: Path | str, anio_corte: int | None = None) -> ParseResult:
    """Filas largas de un XLSX en disco; el anio sale del nombre si no se pasa."""
    ruta = Path(path)
    return parse_bytes(ruta.read_bytes(), anio_corte or anio_from_name(ruta.name))


def xlsx_from_zip(data: bytes) -> tuple[str, bytes]:
    """Nombre y contenido del unico XLSX del ZIP anual publicado por la Secretaria."""
    with zipfile.ZipFile(io.BytesIO(data)) as archivo:
        planillas = [name for name in archivo.namelist() if name.lower().endswith(".xlsx")]
        if len(planillas) != 1:
            raise ValueError(f"el ZIP trae {len(planillas)} XLSX, se esperaba exactamente uno")
        return planillas[0], archivo.read(planillas[0])


def parse_zip(data: bytes, anio_corte: int | None = None) -> ParseResult:
    """Filas largas del ZIP tal como quedo en landing; el anio sale del nombre del XLSX."""
    nombre, planilla = xlsx_from_zip(data)
    return parse_bytes(planilla, anio_corte or anio_from_name(nombre))
